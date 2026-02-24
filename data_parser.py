"""
data_parser.py — Parse S1.xlsx pharma sales data and upload to MongoDB
=======================================================================

Excel structure (S1.xlsx):
  Row 1  : Headers — RANK | UNIT WISE (36 MONTHS) | DATE | PRICE | <36 monthly dates>
  Col  5–40 : Monthly date columns (newest → oldest: Nov-2025 → Dec-2022)
              Some dates have offset days (e.g. 2022-12-22) — normalized to 1st of month

Row types (rows 2+):
  TYPE 1  CATEGORY : rank=None, price=None, date blank/space  → sets current_category
  TYPE 2  BRAND    : rank=int,  price=None                    → sets current_brand / company
  TYPE 3  PRODUCT  : rank=None, price=numeric                 → actual data row

Usage:
  python data_parser.py --file S1.xlsx --mongo "mongodb+srv://..."
  python data_parser.py --file S1.xlsx --export-csv
  python data_parser.py --file S1.xlsx --mongo "..." --export-csv
  python data_parser.py --file S1.xlsx --mongo "..." --db pharma_db
"""

import argparse
import logging
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl

# ─── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# 1. PARSING
# ══════════════════════════════════════════════════════════════════════════════

def normalize_date(dt) -> Optional[datetime]:
    """Normalize any date to the 1st of its month."""
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    try:
        parsed = pd.to_datetime(dt)
        return parsed.replace(day=1).to_pydatetime()
    except Exception:
        return None


def parse_brand_company(raw: str) -> tuple[str, str]:
    """
    Split 'CARICEF            SAM' → ('CARICEF', 'SAM')
    Last whitespace-separated token = company code.
    Everything before = brand name (stripped).
    """
    raw = str(raw).strip()
    parts = raw.split()
    if len(parts) >= 2:
        company = parts[-1].strip()
        brand   = " ".join(parts[:-1]).strip()
    else:
        brand   = raw
        company = "UNKNOWN"
    return brand, company


def parse_excel(filepath: str) -> pd.DataFrame:
    """
    Parse S1.xlsx into a flat long-format DataFrame.

    Returns DataFrame with columns:
      product_category, brand, company, product_name, launch_date,
      price, date, month_label, year, month, units_sold
    """
    log.info(f"Opening: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    log.info(f"Sheet dimensions: {max_row} rows × {max_col} cols")

    # ── Read header row (row 1) ────────────────────────────────────────────
    header_row = ws[1]
    # Columns 5 onward are monthly date columns (0-indexed: 4 onward)
    monthly_dates = []
    monthly_labels = []
    for col_idx in range(4, max_col):          # 0-indexed positions in tuple
        cell_val = header_row[col_idx].value
        norm = normalize_date(cell_val)
        if norm:
            monthly_dates.append(norm)
            monthly_labels.append(norm.strftime("%b-%Y"))   # e.g. "Nov-2025"
        else:
            monthly_dates.append(None)
            monthly_labels.append(None)

    n_months = len(monthly_dates)
    log.info(
        f"Monthly columns: {n_months}  "
        f"({monthly_labels[-1] if monthly_labels else '?'} -> "
        f"{monthly_labels[0] if monthly_labels else '?'})"
    )

    # ── State machine ──────────────────────────────────────────────────────
    current_category: str = "UNKNOWN"
    current_brand:    str = "UNKNOWN"
    current_company:  str = "UNKNOWN"

    records = []

    stats = {"category_rows": 0, "brand_rows": 0, "product_rows": 0, "skipped": 0}

    for row_idx in range(2, max_row + 1):
        row = ws[row_idx]

        rank  = row[0].value
        name  = row[1].value
        date  = row[2].value
        price = row[3].value

        # Skip completely empty rows
        if name is None and rank is None and price is None:
            stats["skipped"] += 1
            continue

        name_str = str(name).strip() if name is not None else ""

        # ── TYPE 1: CATEGORY row ──────────────────────────────────────────
        # rank=None, price=None, date is blank/space/None
        if rank is None and price is None:
            date_str = str(date).strip() if date is not None else ""
            if date_str in ("", " ", "None") or date is None:
                current_category = name_str.upper()
                current_brand    = "UNKNOWN"
                current_company  = "UNKNOWN"
                stats["category_rows"] += 1
                log.info(f"  ► Category: {current_category}  (row {row_idx})")
                continue

        # ── TYPE 2: BRAND row ─────────────────────────────────────────────
        # rank=integer, price=None
        if isinstance(rank, (int, float)) and rank == int(rank) and price is None:
            current_brand, current_company = parse_brand_company(name_str)
            stats["brand_rows"] += 1
            continue

        # ── TYPE 3: PRODUCT row ───────────────────────────────────────────
        # rank=None, price has a numeric value
        if price is not None:
            try:
                price_float = float(price)
            except (ValueError, TypeError):
                stats["skipped"] += 1
                continue

            launch_dt    = normalize_date(date)
            launch_str   = launch_dt.strftime("%Y-%m-%d") if launch_dt else None

            # Read 36 monthly unit values
            for m_idx, (m_date, m_label) in enumerate(zip(monthly_dates, monthly_labels)):
                if m_date is None:
                    continue

                col_pos = 4 + m_idx   # 0-indexed position in row tuple
                if col_pos >= len(row):
                    units = 0
                else:
                    raw_units = row[col_pos].value
                    try:
                        units = int(float(raw_units)) if raw_units is not None else 0
                    except (ValueError, TypeError):
                        units = 0

                records.append({
                    "product_category" : current_category,
                    "brand"            : current_brand,
                    "company"          : current_company,
                    "product_name"     : name_str,
                    "launch_date"      : launch_str,
                    "price"            : price_float,
                    "date"             : m_date.strftime("%Y-%m-%d"),
                    "month_label"      : m_label,
                    "year"             : m_date.year,
                    "month"            : m_date.month,
                    "units_sold"       : units,
                })
            stats["product_rows"] += 1
            continue

        # Anything else: skip
        stats["skipped"] += 1

    log.info(
        f"Parsing complete — "
        f"categories={stats['category_rows']}, "
        f"brands={stats['brand_rows']}, "
        f"products={stats['product_rows']}, "
        f"skipped={stats['skipped']}"
    )
    log.info(f"Total records generated: {len(records):,}")

    df = pd.DataFrame(records)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 2. PRODUCT METADATA SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def build_product_metadata(df: pd.DataFrame) -> pd.DataFrame:
    """
    One row per unique product_name with aggregate metadata.
    """
    agg = (
        df.groupby(
            ["product_name", "product_category", "brand", "company",
             "price", "launch_date"],
            as_index=False,
        )
        .agg(
            first_sale  = ("date", "min"),
            last_sale   = ("date", "max"),
            total_units = ("units_sold", "sum"),
            months_data = ("date", "count"),
        )
        .sort_values(["product_category", "total_units"], ascending=[True, False])
    )
    return agg


# ══════════════════════════════════════════════════════════════════════════════
# 3. MONGODB UPLOAD
# ══════════════════════════════════════════════════════════════════════════════

def upload_to_mongo(df: pd.DataFrame, meta_df: pd.DataFrame, mongo_url: str, db_name: str):
    """
    Upload parsed data to MongoDB:
      - sales_data    : full long-format records (drop + re-insert)
      - products      : product metadata (drop + re-insert)
    """
    try:
        from pymongo import MongoClient, ASCENDING
    except ImportError:
        log.error("pymongo not installed. Run: pip install pymongo")
        sys.exit(1)

    log.info(f"Connecting to MongoDB …")
    client = MongoClient(mongo_url, serverSelectionTimeoutMS=10_000)
    # Verify connection
    client.admin.command("ping")
    db = client[db_name]
    log.info(f"Connected to database: '{db_name}'")

    # ── sales_data ─────────────────────────────────────────────────────────
    col_sales = db["sales_data"]
    log.info("Dropping existing sales_data …")
    col_sales.drop()

    records = df.to_dict(orient="records")
    total   = len(records)
    batch   = 5_000
    inserted = 0

    log.info(f"Uploading {total:,} records in batches of {batch:,} …")
    for start in range(0, total, batch):
        chunk = records[start : start + batch]
        col_sales.insert_many(chunk, ordered=False)
        inserted += len(chunk)
        pct = inserted / total * 100
        log.info(f"  sales_data: {inserted:>7,} / {total:,}  ({pct:.1f}%)")

    # ── Indexes on sales_data ─────────────────────────────────────────────
    log.info("Creating indexes on sales_data …")
    for field in ["product_category", "product_name", "company", "date", "year"]:
        col_sales.create_index([(field, ASCENDING)])
    col_sales.create_index([("product_name", ASCENDING), ("date", ASCENDING)])
    col_sales.create_index([("product_category", ASCENDING), ("year", ASCENDING)])
    log.info("Indexes created.")

    # ── products (metadata) ────────────────────────────────────────────────
    col_meta = db["products"]
    log.info("Dropping existing products collection …")
    col_meta.drop()

    meta_records = meta_df.to_dict(orient="records")
    col_meta.insert_many(meta_records, ordered=False)
    col_meta.create_index([("product_name", ASCENDING)], unique=True)
    col_meta.create_index([("product_category", ASCENDING)])
    col_meta.create_index([("company", ASCENDING)])
    log.info(f"products collection: {len(meta_records):,} documents inserted.")

    client.close()
    log.info("MongoDB upload complete. Connection closed.")


# ══════════════════════════════════════════════════════════════════════════════
# 4. CSV EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def export_csv(df: pd.DataFrame, meta_df: pd.DataFrame, output_dir: str = "."):
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    sales_path = out / "sales_data.csv"
    meta_path  = out / "products_metadata.csv"

    df.to_csv(sales_path, index=False)
    meta_df.to_csv(meta_path, index=False)

    log.info(f"CSV exported -> {sales_path}  ({len(df):,} rows)")
    log.info(f"CSV exported -> {meta_path}  ({len(meta_df):,} rows)")


# ══════════════════════════════════════════════════════════════════════════════
# 5. DIAGNOSTICS / SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def print_summary(df: pd.DataFrame, meta_df: pd.DataFrame):
    print()
    print("=" * 62)
    print("  PARSE SUMMARY")
    print("=" * 62)
    print(f"  Total records      : {len(df):>10,}")
    print(f"  Unique products    : {df['product_name'].nunique():>10,}")
    print(f"  Unique companies   : {df['company'].nunique():>10,}")
    print(f"  Unique categories  : {df['product_category'].nunique():>10,}")
    print(f"  Date range         : {df['date'].min()}  ->  {df['date'].max()}")
    print()
    print("  Records by category:")
    cat_counts = df.groupby("product_category")["units_sold"].agg(
        Records="count", Total_Units="sum", Products=lambda x: x.index.nunique()
    )

    # simpler groupby
    for cat, grp in df.groupby("product_category"):
        print(
            f"    {cat:<22}  "
            f"records={len(grp):>7,}  "
            f"products={grp['product_name'].nunique():>4}  "
            f"total_units={grp['units_sold'].sum():>12,}"
        )
    print()
    print("  Date columns (months):")
    months = sorted(df["date"].unique())
    print(f"    {months[0]}  ->  {months[-1]}  ({len(months)} months)")
    print("=" * 62)
    print()


# ══════════════════════════════════════════════════════════════════════════════
# 6. CLI ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Parse S1.xlsx pharma sales data and upload to MongoDB",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Upload to MongoDB
  python data_parser.py --file S1.xlsx --mongo "mongodb+srv://user:pass@cluster/pharma_db"

  # Export to CSV only
  python data_parser.py --file S1.xlsx --export-csv

  # Both upload and export
  python data_parser.py --file S1.xlsx --mongo "..." --export-csv

  # Custom database name
  python data_parser.py --file S1.xlsx --mongo "..." --db my_pharma_db
        """,
    )
    parser.add_argument(
        "--file", "-f",
        default="S1.xlsx",
        help="Path to the Excel file (default: S1.xlsx)",
    )
    parser.add_argument(
        "--mongo", "-m",
        default=None,
        help="MongoDB connection URL. If omitted, reads MONGODB_URL from .env",
    )
    parser.add_argument(
        "--db",
        default="pharma_db",
        help="MongoDB database name (default: pharma_db)",
    )
    parser.add_argument(
        "--export-csv",
        action="store_true",
        help="Also export parsed data to CSV files in current directory",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory for CSV export (default: current directory)",
    )
    args = parser.parse_args()

    # ── Resolve file path ──────────────────────────────────────────────────
    file_path = Path(args.file)
    if not file_path.exists():
        # Try common locations
        fallbacks = [
            Path("../S1.xlsx"),
            Path(r"C:\Users\Expo\Desktop\Fi\S1.xlsx"),
        ]
        for fb in fallbacks:
            if fb.exists():
                file_path = fb
                log.info(f"Found file at: {file_path}")
                break
        else:
            log.error(f"File not found: {args.file}")
            log.error("Provide correct path with --file /path/to/S1.xlsx")
            sys.exit(1)

    # ── Resolve MongoDB URL ────────────────────────────────────────────────
    mongo_url = args.mongo
    if not mongo_url:
        try:
            from dotenv import load_dotenv
            import os
            load_dotenv()
            mongo_url = os.getenv("MONGODB_URL")
        except ImportError:
            pass

    t0 = time.time()

    # ── Parse ──────────────────────────────────────────────────────────────
    df = parse_excel(str(file_path))

    if df.empty:
        log.error("No records parsed. Check file structure.")
        sys.exit(1)

    meta_df = build_product_metadata(df)

    print_summary(df, meta_df)

    # ── Export CSV ─────────────────────────────────────────────────────────
    if args.export_csv:
        export_csv(df, meta_df, args.output_dir)

    # ── Upload to MongoDB ──────────────────────────────────────────────────
    if mongo_url:
        upload_to_mongo(df, meta_df, mongo_url, args.db)
    else:
        log.warning(
            "No MongoDB URL provided. Use --mongo or set MONGODB_URL in .env. "
            "Data was parsed but NOT uploaded."
        )

    elapsed = time.time() - t0
    log.info(f"Total time: {elapsed:.1f}s")


if __name__ == "__main__":
    main()
